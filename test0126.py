import openpyxl
import io
import sys

# 改编标准输出默认编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
path = "./file/安博通审计产品线目录树管理V2.1【20201218】.xlsx"
# 打开一个工作薄 workbook
wb = openpyxl.load_workbook(path)
# 一个工作薄中含有多个工作表 wooksheet，sheetnames 返回工作表名称
ws = wb.sheetnames
sheet = wb['审计产品线相关产品目录树']
# print(sheet['A2'])
# print(sheet.title)
# print(sheet.max_row)
# print(sheet.min_row)
# print(sheet.max_column)
# print(list(sheet.values))
# print(sheet.iter_rows())
# content = list(sheet.values)
input('请输入：')
content = sheet.values
data = []
for i in content:
    if i[7] == '√' and isinstance(i[6], int):
        data.append(i[6])
print(data)
